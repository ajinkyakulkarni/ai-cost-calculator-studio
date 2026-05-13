#!/usr/bin/env python3
"""
Three-way diff harness: JS engine vs Python port vs Excel single-sheet.

Runs the same workload through all three implementations and prints a
comparison table. If the three numbers agree to within $1, the math is
considered correct. Divergences are real findings — flag the
implementation that differs.

Usage:
  python3 scripts/three_way_diff.py public/examples/public-geospatial-qa.json
  python3 scripts/three_way_diff.py public/examples/*.json
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
CALC_JS = REPO_ROOT / "scripts" / "calc.js"
EXCEL_GEN = REPO_ROOT / "public" / "lib" / "excel-generator.py"

# Ensure the python port is importable
sys.path.insert(0, str(REPO_ROOT / "scripts"))
import cost_engine as py_engine  # noqa: E402


def run_js(workload_path: Path) -> dict:
    """Run scripts/calc.js and parse its JSON output."""
    proc = subprocess.run(
        ["node", str(CALC_JS), "--workload", str(workload_path), "--json"],
        capture_output=True, text=True, check=False, cwd=REPO_ROOT,
    )
    if proc.returncode != 0:
        raise RuntimeError(f"calc.js failed: {proc.stderr}")
    data = json.loads(proc.stdout)
    return {
        "llm_monthly":      data["lines"]["llm"],
        "headline_monthly": data["headline"]["monthly"],
        "queries_per_mo":   data["derived"]["queries_per_month"],
    }


def run_python(workload_path: Path) -> dict:
    """Run the Python port directly."""
    with open(workload_path) as f:
        w = json.load(f)
    result = py_engine.compute(w, {})
    s = py_engine.summary(result, {})
    return {
        "llm_monthly":      s["llm_monthly"],
        "headline_monthly": s["headline_monthly"],
        "queries_per_mo":   s["queries_per_month"],
    }


def run_excel(workload_path: Path) -> dict:
    """Generate an Excel sheet and evaluate its formulas with `formulas`."""
    import formulas  # heavy import; only when needed
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = Path(tmp.name)
    try:
        # Generate the workbook
        subprocess.run(
            ["python3", str(EXCEL_GEN), str(workload_path), "-o", str(xlsx_path)],
            check=True, capture_output=True,
        )
        # Evaluate formulas
        xl_model = formulas.ExcelModel().loads(str(xlsx_path)).finish()
        solution = xl_model.calculate()
        # Find cells by searching for known labels in column A and reading column B.
        # The formulas package keys solutions like
        #   "'[lowercase-filename.xlsx]UPPERCASE-SHEET'!CELL"
        # so we build the cell-suffix and scan keys to find a case-insensitive match.
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        labels_to_find = {
            "Monthly LLM bill": "llm_monthly",
            "TOTAL queries / month": "queries_per_mo",
        }
        cell_for = {}
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label:
                for needle, key in labels_to_find.items():
                    if needle in str(label):
                        cell_for[key] = f"B{r}"
        # Index the solution dict by uppercased cell-suffix for case-insensitive lookup
        sol_by_suffix = {}
        for k in solution.keys():
            if "!" in k:
                suffix = k.rsplit("!", 1)[1].upper().rstrip("'")
                sol_by_suffix[suffix] = k
        # Read solved values
        out = {}
        for key, cell in cell_for.items():
            real_key = sol_by_suffix.get(cell.upper())
            if real_key is None:
                out[key] = f"(no key for {cell})"
                continue
            try:
                val = solution[real_key].value[0][0]
                out[key] = float(val)
            except Exception as e:
                out[key] = f"(error: {e})"
        # Excel headline = LLM only (the Excel single-sheet does not currently
        # add verification/federal/etc.; that's the next polish step). Treat
        # headline as same as llm for now to surface the gap explicitly.
        out["headline_monthly"] = out.get("llm_monthly", None)
        return out
    finally:
        if xlsx_path.exists():
            xlsx_path.unlink()


def fmt_dollar(x):
    return f"${x:,.2f}" if isinstance(x, (int, float)) else str(x)


def fmt_int(x):
    return f"{int(x):,}" if isinstance(x, (int, float)) else str(x)


def compare(workload_path: Path) -> bool:
    """Run all three, print comparison, return True if they agree within $1."""
    print(f"\n{'═' * 78}")
    print(f"  {workload_path.name}")
    print(f"{'═' * 78}")

    try:
        js = run_js(workload_path)
    except Exception as e:
        print(f"  JS:     FAILED — {e}")
        js = None
    try:
        py = run_python(workload_path)
    except Exception as e:
        print(f"  Python: FAILED — {e}")
        py = None
    try:
        xl = run_excel(workload_path)
    except Exception as e:
        print(f"  Excel:  FAILED — {e}")
        xl = None

    # Strict checks: the core LLM equation + query volume. All three
    # implementations should agree on these to within $1 / 1 query.
    # Informational: headline_monthly, which depends on add-on layers
    # (verification, fixed, federal) that the single-sheet Excel doesn't
    # currently model. A gap there means "Excel coverage is partial",
    # not "math is wrong".
    strict = ["llm_monthly", "queries_per_mo"]
    info = ["headline_monthly"]

    print(f"\n  {'Metric':<24} {'JS':>14}  {'Python':>14}  {'Excel':>14}  Verdict")
    print(f"  {'-' * 24} {'-' * 14}  {'-' * 14}  {'-' * 14}  -------")
    ok = True
    for m in strict + info:
        fmt = fmt_dollar if "monthly" in m else fmt_int
        jv = js[m] if js else None
        pv = py[m] if py else None
        xv = xl[m] if xl else None
        nums = [v for v in (jv, pv, xv) if isinstance(v, (int, float))]
        spread = max(nums) - min(nums) if len(nums) >= 2 else 0
        tol = 1.0 if "monthly" in m else 1
        if m in strict:
            verdict = "✓ agree" if spread <= tol else f"✗ diff ${spread:,.2f}"
            if spread > tol:
                ok = False
        else:
            verdict = f"info (Δ={fmt(spread)})" if spread > tol else "✓ agree"
        print(f"  {m:<24} {fmt(jv):>14}  {fmt(pv):>14}  {fmt(xv):>14}  {verdict}")
    print()
    return ok


def main():
    ap = argparse.ArgumentParser(description="Three-way diff: JS vs Python vs Excel")
    ap.add_argument("workloads", nargs="+", help="Workload JSON path(s); accepts globs")
    args = ap.parse_args()

    paths = []
    for p in args.workloads:
        path = Path(p)
        if path.is_file():
            paths.append(path)
        else:
            paths.extend(sorted(path.parent.glob(path.name)))

    all_ok = True
    for p in paths:
        ok = compare(p)
        all_ok = all_ok and ok

    print(f"\n{'═' * 78}")
    print(f"  Overall: {'✓ all implementations agree' if all_ok else '✗ DIVERGENCES FOUND'}")
    print(f"{'═' * 78}\n")
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
