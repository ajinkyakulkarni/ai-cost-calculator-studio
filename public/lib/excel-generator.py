#!/usr/bin/env python3
"""
Cost Calculator Studio — Excel workbook generator.

Builds a sophisticated .xlsx file from a workload specification. The
workbook contains live formulas mirroring the JS cost engine, so a
procurement officer can paste their numbers and get the same answers
they'd see in the interactive calculator.

Sheets produced:
  README        — methodology, links, how to use
  Workload      — input parameters (editable)
  Shapes        — query shape definitions
  Mix           — traffic mix weights
  Segments      — user populations
  Rate Cards    — per-million-token model rates
  Cost Modes    — Optimistic vs Realistic parameters
  Computation   — intermediate values, all driven by formulas
  Output        — the four-row API-vs-self-host comparison
  Sensitivity   — how the headline moves when key inputs change

Usage:
  python3 excel-generator.py examples/public-geospatial-qa.json -o public-geospatial-qa.xlsx
  python3 excel-generator.py examples/generic-startup-chatbot.json -o startup.xlsx

Requires: openpyxl (pip install openpyxl). Standard library only otherwise.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ----- Editorial palette to match the HTML calculator -----
INK = "1A1A1A"
ACCENT = "8B2331"     # oxblood
ACCENT2 = "C87A52"    # warm copper (secondary)
GOOD = "2B7A3D"
BAD = "B8404A"
PAPER = "FBF9F4"
HIGHLIGHT = "F3ECDB"
HEATMAP_LO = "E8F4E8"  # pale green
HEATMAP_MID = "FFF4D6"  # cream
HEATMAP_HI = "F8D7DA"  # pale red
RULE = "D8D2C5"

THIN = Side(border_style="thin", color=RULE)
THICK = Side(border_style="medium", color=INK)


def style_header(cell, level: int = 1):
    """Section/column header styling."""
    if level == 1:
        cell.font = Font(name="Calibri", size=14, bold=True, color=ACCENT)
    elif level == 2:
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    else:
        cell.font = Font(name="Calibri", size=10, bold=True, color="555555")
        cell.alignment = Alignment(horizontal="left", vertical="center")


def style_label(cell):
    cell.font = Font(name="Calibri", size=11, color=INK)


def style_input(cell):
    cell.font = Font(name="Consolas", size=11, color="0E5C8A", bold=True)
    cell.fill = PatternFill(start_color="EAF3F9", end_color="EAF3F9", fill_type="solid")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_formula(cell):
    cell.font = Font(name="Consolas", size=11, color=INK)
    cell.fill = PatternFill(start_color=PAPER, end_color=PAPER, fill_type="solid")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_result(cell):
    cell.font = Font(name="Consolas", size=12, bold=True, color=ACCENT)
    cell.fill = PatternFill(start_color=HIGHLIGHT, end_color=HIGHLIGHT, fill_type="solid")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def autosize(ws, min_w=12, max_w=60):
    for col_idx, col in enumerate(ws.columns, start=1):
        try:
            longest = max(len(str(c.value)) for c in col if c.value is not None)
        except ValueError:
            longest = min_w
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_w, min(max_w, longest + 2))


# ============================================================
# Visual-polish helpers
# ============================================================

def fmt_usd(cell):
    """Apply $#,##0 number format."""
    cell.number_format = '"$"#,##0'


def fmt_usd_cents(cell):
    """Apply $#,##0.00 number format for small per-query values."""
    cell.number_format = '"$"#,##0.00'


def fmt_pct(cell):
    """Apply 0.0% number format."""
    cell.number_format = "0.0%"


def fmt_num(cell):
    """Apply #,##0 number format for query counts."""
    cell.number_format = "#,##0"


def freeze_below(ws, row, col=1):
    """Freeze rows above `row` and columns left of `col`."""
    ws.freeze_panes = ws.cell(row=row, column=col).coordinate


def heatmap_range(ws, range_str):
    """3-color scale heat-map (green → cream → red, lo → mid → hi)."""
    rule = ColorScaleRule(
        start_type='min', start_color=HEATMAP_LO,
        mid_type='percentile', mid_value=50, mid_color=HEATMAP_MID,
        end_type='max', end_color=HEATMAP_HI,
    )
    ws.conditional_formatting.add(range_str, rule)


def title_banner(ws, title: str, subtitle: str = "", cols: int = 6):
    """Top-of-sheet branded banner. Row 1 = title in ACCENT; row 2 = subtitle in muted."""
    t = ws["A1"]
    t.value = title
    t.font = Font(name="Calibri", size=15, bold=True, color=ACCENT)
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    if subtitle:
        s = ws["A2"]
        s.value = subtitle
        s.font = Font(name="Calibri", size=10, italic=True, color="666666")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)


def add_bar_chart(ws, *, title: str, data_ref: str, cats_ref: str, anchor: str, height: int = 9, width: int = 16):
    """Drop a horizontal-bar chart at `anchor`."""
    ch = BarChart()
    ch.type = "bar"
    ch.style = 11
    ch.title = title
    ch.y_axis.title = None
    ch.x_axis.title = "USD / month"
    ch.legend = None
    data = Reference(ws, range_string=data_ref)
    cats = Reference(ws, range_string=cats_ref)
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(cats)
    ch.height = height
    ch.width = width
    ch.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(ch, anchor)


def add_line_chart(ws, *, title: str, data_ref: str, cats_ref: str, anchor: str, height: int = 9, width: int = 18):
    """Drop a line chart at `anchor` (used for sensitivity sweeps)."""
    ch = LineChart()
    ch.style = 10
    ch.title = title
    ch.y_axis.title = "USD / month"
    ch.x_axis.title = "Sensitivity (% of baseline)"
    data = Reference(ws, range_string=data_ref)
    cats = Reference(ws, range_string=cats_ref)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.height = height
    ch.width = width
    ws.add_chart(ch, anchor)


# ============================================================
# Per-sheet builders
# ============================================================

def build_readme(ws, w: dict):
    ws.title = "README"

    title = ws["A1"]
    title.value = f"Cost Calculator · {w['deployment'].get('name', 'AI Agent')}"
    style_header(title, 1)
    ws["A1"].alignment = Alignment(vertical="center")

    ws["A2"] = w["deployment"].get("agency", "")
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="555555")

    ws["A4"] = "About this workbook"
    style_header(ws["A4"], 2)

    rows = [
        "",
        "This Excel workbook is a parameterized cost model for the deployment named above.",
        "All numbers in the Output sheet are driven by formulas referencing the Workload,",
        "Shapes, Mix, Segments, Rate Cards, and Cost Modes sheets.",
        "",
        "How to use:",
        "  1. Open the Workload sheet and edit any blue-highlighted input cell.",
        "  2. The Computation and Output sheets recalculate automatically.",
        "  3. The Sensitivity sheet shows how the headline number moves when key",
        "     inputs swing ±30%.",
        "",
        "Methodology:",
        "  Same arithmetic as the Cost Calculator Studio web app and the companion paper.",
        "  See lib/cost-engine.js for the canonical implementation.",
        "  See docs/papers/arxiv/paper.tex for the academic write-up.",
        "",
        "Cite as:",
        "  Kulkarni, A. & Parajuli, P. (2026). Cost Modeling for Public-Facing",
        "  LLM Chat Applications: An Equal-Budget, Refusal-Aware Comparison",
        "  of Commercial APIs and Self-Hosted GPU Fleets.",
        "",
        "Schema version: " + w.get("schemaVersion", "1.0"),
    ]

    for i, text in enumerate(rows, start=5):
        ws[f"A{i}"] = text
        style_label(ws[f"A{i}"])

    ws.column_dimensions["A"].width = 95
    ws.row_dimensions[1].height = 30


def build_workload(ws, w: dict):
    ws.title = "Workload"

    ws["A1"] = "Workload Specification"
    style_header(ws["A1"], 1)
    ws.merge_cells("A1:C1")

    section_offsets = []

    row = 3
    ws[f"A{row}"] = "Deployment"; style_header(ws[f"A{row}"], 2); ws.merge_cells(f"A{row}:C{row}")
    row += 1
    deployment_fields = [
        ("name", "Name"),
        ("agency", "Agency"),
        ("description", "Description"),
        ("publicFacing", "Public-facing"),
        ("fedrampTier", "FedRAMP tier"),
    ]
    for key, label in deployment_fields:
        ws[f"A{row}"] = label; style_label(ws[f"A{row}"])
        v = w["deployment"].get(key, "")
        ws[f"B{row}"] = v if not isinstance(v, bool) else ("yes" if v else "no")
        style_input(ws[f"B{row}"])
        row += 1

    row += 1
    ws[f"A{row}"] = "Anchor query"; style_header(ws[f"A{row}"], 2); ws.merge_cells(f"A{row}:C{row}")
    row += 1
    aq = w["anchor_query"]
    anchor_fields = [
        ("input_tokens", "Input tokens (anchor)"),
        ("output_tokens", "Output tokens (anchor)"),
        ("cache_rate_baseline", "Cache rate baseline (0–1)"),
        ("session_baseline_turns", "Baseline session turns"),
    ]
    anchor_input_cell = {}
    for key, label in anchor_fields:
        ws[f"A{row}"] = label; style_label(ws[f"A{row}"])
        ws[f"B{row}"] = aq.get(key, 0)
        style_input(ws[f"B{row}"])
        anchor_input_cell[key] = f"Workload!B{row}"
        row += 1

    row += 1
    ws[f"A{row}"] = "Daily spend cap"; style_header(ws[f"A{row}"], 2); ws.merge_cells(f"A{row}:C{row}")
    row += 1
    cap = w.get("daily_cap", {"enabled": True, "amount_usd": 1500, "burst_days": 7, "burst_factor": 1.0})
    cap_fields = [("enabled", "Enabled (yes/no)"), ("amount_usd", "$/day"), ("burst_days", "Burst days/month"), ("burst_factor", "Burst factor")]
    cap_cell = {}
    for key, label in cap_fields:
        ws[f"A{row}"] = label; style_label(ws[f"A{row}"])
        v = cap.get(key, "")
        ws[f"B{row}"] = ("yes" if v else "no") if isinstance(v, bool) else v
        style_input(ws[f"B{row}"])
        cap_cell[key] = f"Workload!B{row}"
        row += 1

    row += 1
    ws[f"A{row}"] = "Bot factor"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = 1.5
    style_input(ws[f"B{row}"])
    bot_factor_cell = f"Workload!B{row}"
    row += 1

    autosize(ws, min_w=20, max_w=70)
    return {
        "anchor_input": anchor_input_cell,
        "cap": cap_cell,
        "bot_factor": bot_factor_cell,
    }


def build_shapes(ws, w: dict):
    ws.title = "Shapes"
    headers = ["Shape", "Input factor", "Output factor", "Cache eligible", "Description"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h); style_header(c, 2)

    row = 2
    shape_rows = {}
    for name, shape in w["shapes"].items():
        ws.cell(row=row, column=1, value=name).font = Font(name="Consolas", bold=True, color=ACCENT)
        ws.cell(row=row, column=2, value=shape["input_factor"]); style_input(ws.cell(row=row, column=2))
        ws.cell(row=row, column=3, value=shape["output_factor"]); style_input(ws.cell(row=row, column=3))
        ws.cell(row=row, column=4, value="yes" if shape["cache_eligible"] else "no"); style_input(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5, value=shape.get("description", ""))
        shape_rows[name] = row
        row += 1
    autosize(ws)
    return shape_rows


def build_mix(ws, w: dict):
    ws.title = "Mix"
    shapes = list(w["shapes"].keys())
    headers = ["Mix", "Label"] + shapes + ["Sum"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h); style_header(c, 2)

    row = 2
    mix_rows = {}
    for mix_name, mix in w["mix"].items():
        ws.cell(row=row, column=1, value=mix_name).font = Font(name="Consolas", bold=True, color=ACCENT)
        ws.cell(row=row, column=2, value=mix.get("label", mix_name))
        for i, shape in enumerate(shapes, start=3):
            ws.cell(row=row, column=i, value=mix.get("weights", {}).get(shape, 0))
            style_input(ws.cell(row=row, column=i))
        # Sum formula at the end
        first_col = get_column_letter(3)
        last_col = get_column_letter(2 + len(shapes))
        ws.cell(row=row, column=3 + len(shapes), value=f"=SUM({first_col}{row}:{last_col}{row})")
        style_formula(ws.cell(row=row, column=3 + len(shapes)))
        mix_rows[mix_name] = row
        row += 1
    autosize(ws)
    return mix_rows


def build_segments(ws, w: dict):
    ws.title = "Segments"
    headers = ["ID", "Label", "MAU", "Sessions/day", "Q/session", "Bot factor?"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h); style_header(c, 2)

    row = 2
    seg_rows = {}
    for seg in w["segments"]:
        ws.cell(row=row, column=1, value=seg["id"]).font = Font(name="Consolas", bold=True, color=ACCENT)
        ws.cell(row=row, column=2, value=seg.get("label", seg["id"]))
        ws.cell(row=row, column=3, value=seg["mau"]); style_input(ws.cell(row=row, column=3))
        ws.cell(row=row, column=4, value=seg["sessions_per_day"]); style_input(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5, value=seg["questions_per_session"]); style_input(ws.cell(row=row, column=5))
        ws.cell(row=row, column=6, value="yes" if seg.get("applyBotFactor") else "no"); style_input(ws.cell(row=row, column=6))
        seg_rows[seg["id"]] = row
        row += 1
    autosize(ws)
    return seg_rows


def build_rate_cards(ws, w: dict):
    ws.title = "Rate Cards"
    # Default rate cards (mirrors lib/cost-engine.js)
    DEFAULT = {
        'gpt-5.5':         (5.00, 0.50,  30.00),
        'gpt-5.4':         (2.50, 0.25,  15.00),
        'gpt-5.2':         (1.75, 0.175, 14.00),
        'gpt-5.1':         (1.25, 0.125, 10.00),
        'gpt-5':           (1.25, 0.125, 10.00),
        'gpt-5-mini':      (0.25, 0.025,  2.00),
        'gpt-5-nano':      (0.05, 0.005,  0.40),
        'gpt-4o':          (2.50, 1.25,  10.00),
        'gpt-4o-mini':     (0.15, 0.075,  0.60),
        'claude-opus-4.7': (5.00, 0.50,  25.00),
        'gemini-3.1-pro':  (2.00, 0.20,  12.00),
    }
    rates = dict(DEFAULT)
    for k, v in (w.get("rate_cards") or {}).items():
        rates[k] = (v.get("input_per_million"), v.get("cached_per_million"), v.get("output_per_million"))
    headers = ["Model", "Input ($/M)", "Cached ($/M)", "Output ($/M)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h); style_header(c, 2)
    row = 2
    rate_rows = {}
    for name, (i, c, o) in rates.items():
        ws.cell(row=row, column=1, value=name).font = Font(name="Consolas", bold=True, color=ACCENT)
        ws.cell(row=row, column=2, value=i); style_input(ws.cell(row=row, column=2))
        ws.cell(row=row, column=3, value=c); style_input(ws.cell(row=row, column=3))
        ws.cell(row=row, column=4, value=o); style_input(ws.cell(row=row, column=4))
        rate_rows[name] = row
        row += 1
    autosize(ws)
    return rate_rows


def build_cost_modes(ws):
    ws.title = "Cost Modes"
    DEFAULTS = {
        "optimistic": {"ops_monthly": 350,  "fte_monthly": 2500, "setup_amortized":    0, "throughput_derate": 1.00, "discount_1yr": 0.40, "discount_3yr": 0.60},
        "realistic":  {"ops_monthly": 1800, "fte_monthly": 8000, "setup_amortized": 8333, "throughput_derate": 0.75, "discount_1yr": 0.33, "discount_3yr": 0.55},
    }
    headers = ["Parameter", "Optimistic", "Realistic"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h); style_header(c, 2)
    rows = list(DEFAULTS["optimistic"].keys())
    for i, key in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=key.replace("_", " ").title())
        ws.cell(row=i, column=2, value=DEFAULTS["optimistic"][key]); style_input(ws.cell(row=i, column=2))
        ws.cell(row=i, column=3, value=DEFAULTS["realistic"][key]); style_input(ws.cell(row=i, column=3))
    autosize(ws)


def build_computation(ws, w: dict, refs: dict):
    """The intermediate values: queries, blended cost, monthly gross, capped, refused."""
    ws.title = "Computation"

    ws["A1"] = "Computation"
    style_header(ws["A1"], 1)
    ws.merge_cells("A1:C1")

    row = 3
    ws[f"A{row}"] = "Selections (change here to model)"; style_header(ws[f"A{row}"], 2); ws.merge_cells(f"A{row}:C{row}")
    row += 1
    default_model = w["defaults"]["model"]
    default_mix = w["defaults"]["mix"]
    default_tier = w["defaults"]["tier"]

    ws[f"A{row}"] = "Model"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = default_model; style_input(ws[f"B{row}"])
    model_cell = f"Computation!B{row}"
    row += 1

    ws[f"A{row}"] = "Mix"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = default_mix; style_input(ws[f"B{row}"])
    mix_cell = f"Computation!B{row}"
    row += 1

    ws[f"A{row}"] = "Tier"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = default_tier; style_input(ws[f"B{row}"])
    tier_cell = f"Computation!B{row}"
    row += 1

    ws[f"A{row}"] = "Cost mode"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = w["defaults"].get("cost_mode", "optimistic"); style_input(ws[f"B{row}"])
    cost_mode_cell = f"Computation!B{row}"
    row += 1

    # Tier multiplier lookup
    row += 1
    ws[f"A{row}"] = "Tier multiplier"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f'=IF({tier_cell}="standard",1,IF({tier_cell}="flex",0.5,IF({tier_cell}="batch",0.5,IF({tier_cell}="priority",2.5,1))))'
    style_formula(ws[f"B{row}"])
    tier_mult_cell = f"Computation!B{row}"
    row += 1

    # Per-model rate lookup via VLOOKUP
    rate_rows = refs["rate_rows"]
    n_rates = len(rate_rows)
    rates_range_in = f"'Rate Cards'!A2:B{1 + n_rates}"
    rates_range_cached = f"'Rate Cards'!A2:C{1 + n_rates}"
    rates_range_out = f"'Rate Cards'!A2:D{1 + n_rates}"

    ws[f"A{row}"] = "Rate (input $/M)"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f'=VLOOKUP({model_cell},{rates_range_in},2,FALSE)'; style_formula(ws[f"B{row}"])
    p_in_cell = f"Computation!B{row}"
    row += 1
    ws[f"A{row}"] = "Rate (cached $/M)"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f'=VLOOKUP({model_cell},{rates_range_cached},3,FALSE)'; style_formula(ws[f"B{row}"])
    p_cached_cell = f"Computation!B{row}"
    row += 1
    ws[f"A{row}"] = "Rate (output $/M)"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f'=VLOOKUP({model_cell},{rates_range_out},4,FALSE)'; style_formula(ws[f"B{row}"])
    p_out_cell = f"Computation!B{row}"
    row += 1

    row += 1
    ws[f"A{row}"] = "Per-shape cost (full pipeline)"; style_header(ws[f"A{row}"], 2); ws.merge_cells(f"A{row}:C{row}")
    row += 1

    anchor_in_cell = refs["anchor_input"]["input_tokens"]
    anchor_out_cell = refs["anchor_input"]["output_tokens"]
    cache_base_cell = refs["anchor_input"]["cache_rate_baseline"]

    # Compute per-segment effective cache + per-query cost
    ws[f"A{row}"] = "Effective cache (per segment)"; style_header(ws[f"A{row}"], 2); ws.merge_cells(f"A{row}:C{row}")
    row += 1
    ws[f"A{row}"] = "Segment ID"
    ws[f"B{row}"] = "Eff cache rate"
    ws[f"C{row}"] = "Per-query $ (blended)"
    style_header(ws[f"A{row}"], 3); style_header(ws[f"B{row}"], 3); style_header(ws[f"C{row}"], 3)
    row += 1

    seg_rows = refs["seg_rows"]
    seg_eff_cells = {}
    seg_pq_cells = {}
    for seg_id, seg_row in seg_rows.items():
        q_per_session_cell = f"Segments!E{seg_row}"
        ws[f"A{row}"] = seg_id
        # eff cache = clamp(base + (q-6)*0.01, 0.5, 0.94)
        ws[f"B{row}"] = (
            f'=MAX(0.5,MIN(0.94,{cache_base_cell}+({q_per_session_cell}-6)*0.01))'
        )
        style_formula(ws[f"B{row}"])
        seg_eff_cells[seg_id] = f"Computation!B{row}"

        # Per-query blended cost across mix
        # We'll use a simplified formula: compute weighted sum manually for top 5 shapes
        # Use mix lookup per-shape
        mix_row = refs["mix_rows"][default_mix]  # Use default mix; can be parameterized later
        shape_rows = refs["shape_rows"]
        # Build SUMPRODUCT manually
        formula_parts = []
        col_idx = 3  # mix weights start at col C
        for shape_name, shape_row in shape_rows.items():
            mix_w = f'INDEX(Mix!{get_column_letter(col_idx)}:{get_column_letter(col_idx)},{mix_row})'
            in_factor = f"Shapes!B{shape_row}"
            out_factor = f"Shapes!C{shape_row}"
            cache_elig = f"Shapes!D{shape_row}"  # "yes" or "no"
            cache_use = f'IF({cache_elig}="yes",{seg_eff_cells[seg_id]},0)'
            in_tokens = f"({anchor_in_cell}*{in_factor})"
            out_tokens = f"({anchor_out_cell}*{out_factor})"
            cached = f"({in_tokens}*{cache_use})"
            uncached = f"({in_tokens}-{cached})"
            shape_cost = (
                f"({uncached}*{p_in_cell}/1000000"
                f"+{cached}*{p_cached_cell}/1000000"
                f"+{out_tokens}*{p_out_cell}/1000000)"
                f"*{tier_mult_cell}"
            )
            formula_parts.append(f"{mix_w}*{shape_cost}")
            col_idx += 1
        sumprod = "+".join(formula_parts)
        ws[f"C{row}"] = f"={sumprod}"
        style_formula(ws[f"C{row}"])
        seg_pq_cells[seg_id] = f"Computation!C{row}"
        row += 1

    # Monthly query volume per segment
    row += 1
    ws[f"A{row}"] = "Monthly query volume"; style_header(ws[f"A{row}"], 2); ws.merge_cells(f"A{row}:C{row}")
    row += 1
    ws[f"A{row}"] = "Segment ID"; ws[f"B{row}"] = "Queries/mo"; ws[f"C{row}"] = "$ / mo (gross)"
    style_header(ws[f"A{row}"], 3); style_header(ws[f"B{row}"], 3); style_header(ws[f"C{row}"], 3)
    row += 1
    bot_cell = refs["bot_factor"]
    seg_q_cells = {}
    seg_cost_cells = {}
    for seg_id, seg_row in seg_rows.items():
        mau_cell = f"Segments!C{seg_row}"
        spd_cell = f"Segments!D{seg_row}"
        qps_cell = f"Segments!E{seg_row}"
        bot_flag_cell = f"Segments!F{seg_row}"
        beta = f'IF({bot_flag_cell}="yes",{bot_cell},1)'
        ws[f"A{row}"] = seg_id
        ws[f"B{row}"] = f"={mau_cell}*{spd_cell}*30*{qps_cell}*{beta}"
        style_formula(ws[f"B{row}"])
        seg_q_cells[seg_id] = f"Computation!B{row}"
        ws[f"C{row}"] = f"=B{row}*{seg_pq_cells[seg_id]}"
        style_formula(ws[f"C{row}"])
        seg_cost_cells[seg_id] = f"Computation!C{row}"
        row += 1

    # Totals
    row += 1
    ws[f"A{row}"] = "TOTAL queries/mo"; style_label(ws[f"A{row}"])
    total_q_formula = "=" + "+".join(seg_q_cells.values())
    ws[f"B{row}"] = total_q_formula
    style_result(ws[f"B{row}"])
    total_q_cell = f"Computation!B{row}"
    row += 1

    ws[f"A{row}"] = "TOTAL API gross $/mo"; style_label(ws[f"A{row}"])
    total_cost_formula = "=" + "+".join(seg_cost_cells.values())
    ws[f"B{row}"] = total_cost_formula
    style_result(ws[f"B{row}"])
    api_gross_cell = f"Computation!B{row}"
    row += 1

    # Daily cap clamping
    cap_amount_cell = refs["cap"]["amount_usd"]
    cap_burst_days = refs["cap"]["burst_days"]
    ws[f"A{row}"] = "Daily cap budget ($/mo)"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f"={cap_amount_cell}*30"
    style_formula(ws[f"B{row}"])
    cap_monthly_cell = f"Computation!B{row}"
    row += 1

    ws[f"A{row}"] = "API monthly capped"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f"=MIN({api_gross_cell},{cap_monthly_cell})"
    style_result(ws[f"B{row}"])
    api_capped_cell = f"Computation!B{row}"
    row += 1

    ws[f"A{row}"] = "Refused queries"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f"=IF({api_gross_cell}>{api_capped_cell},({api_gross_cell}-{api_capped_cell})/{api_gross_cell}*{total_q_cell},0)"
    style_result(ws[f"B{row}"])
    fmt_num(ws[f"B{row}"])
    refused_cell = f"Computation!B{row}"
    row += 1

    # ------------------------------------------------------------
    # Self-host monthly cost (per cost mode)
    # GPU monthly = hourly × 730 × (1 − discount_3yr)
    # Total       = ops + fte + setup_amort + gpu_monthly
    # Uses a single-replica simplification — full instance-count math
    # (Eq. 6) requires peak-tps × throughput-derate; see cost-engine.js
    # for the production formula. This Excel approximation gives a
    # ballpark for procurement comparison; the live calc gives the
    # exact number for a given workload spec.
    # ------------------------------------------------------------
    row += 1
    ws[f"A{row}"] = "Self-host cost (single replica)"; style_header(ws[f"A{row}"], 2); ws.merge_cells(f"A{row}:C{row}")
    row += 1

    # First GPU option from the workload (typically a g6 or A100 SKU)
    gpu_opts = w.get("self_host", {}).get("gpu_options", {})
    if gpu_opts:
        first_gpu_id, first_gpu = next(iter(gpu_opts.items()))
        gpu_hourly = first_gpu.get("hourly", 8.0)
        gpu_label = first_gpu.get("label", first_gpu_id)
    else:
        gpu_hourly = 8.0
        gpu_label = "g6.12xlarge (default)"

    ws[f"A{row}"] = "GPU SKU"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = gpu_label
    row += 1
    ws[f"A{row}"] = "GPU hourly ($)"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = gpu_hourly; style_input(ws[f"B{row}"])
    fmt_usd_cents(ws[f"B{row}"])
    gpu_hourly_cell = f"Computation!B{row}"
    row += 1

    # Per-mode columns: Optimistic in B, Realistic in C
    ws[f"A{row}"] = ""
    ws[f"B{row}"] = "Optimistic"; style_header(ws[f"B{row}"], 3)
    ws[f"C{row}"] = "Realistic"; style_header(ws[f"C{row}"], 3)
    row += 1

    ws[f"A{row}"] = "Ops + FTE + setup ($/mo)"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = "='Cost Modes'!B2+'Cost Modes'!B3+'Cost Modes'!B4"
    ws[f"C{row}"] = "='Cost Modes'!C2+'Cost Modes'!C3+'Cost Modes'!C4"
    style_formula(ws[f"B{row}"]); style_formula(ws[f"C{row}"]); fmt_usd(ws[f"B{row}"]); fmt_usd(ws[f"C{row}"])
    fixed_opt_cell = f"Computation!B{row}"
    fixed_real_cell = f"Computation!C{row}"
    row += 1

    ws[f"A{row}"] = "GPU monthly ($/mo, 3-yr RI)"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f"={gpu_hourly_cell}*730*(1-'Cost Modes'!B7)"
    ws[f"C{row}"] = f"={gpu_hourly_cell}*730*(1-'Cost Modes'!C7)"
    style_formula(ws[f"B{row}"]); style_formula(ws[f"C{row}"]); fmt_usd(ws[f"B{row}"]); fmt_usd(ws[f"C{row}"])
    gpu_opt_cell = f"Computation!B{row}"
    gpu_real_cell = f"Computation!C{row}"
    row += 1

    ws[f"A{row}"] = "Self-host total ($/mo)"; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = f"={fixed_opt_cell}+{gpu_opt_cell}"
    ws[f"C{row}"] = f"={fixed_real_cell}+{gpu_real_cell}"
    style_result(ws[f"B{row}"]); style_result(ws[f"C{row}"]); fmt_usd(ws[f"B{row}"]); fmt_usd(ws[f"C{row}"])
    sh_opt_cell = f"Computation!B{row}"
    sh_real_cell = f"Computation!C{row}"
    row += 1

    autosize(ws, min_w=22, max_w=70)
    freeze_below(ws, row=4)
    return {
        "model_cell": model_cell,
        "total_q": total_q_cell,
        "api_gross": api_gross_cell,
        "api_capped": api_capped_cell,
        "refused": refused_cell,
        "cost_mode": cost_mode_cell,
        "self_host_opt": sh_opt_cell,
        "self_host_real": sh_real_cell,
    }


def build_output(ws, w: dict, comp_refs: dict, cap_refs: dict):
    ws.title = "Output"
    title_banner(
        ws,
        "Output · Equal-budget comparison",
        f"API capped at ${cap_refs.get('amount_usd', 1500):,}/day vs self-host (Optimistic + Realistic modes) — all four strategies on one table.",
        cols=4,
    )

    headers = ["Strategy", "Monthly cost ($)", "Queries served/mo", "Notes"]
    row = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, 2)
    row += 1

    data_start = row

    # Row 1: API capped
    ws.cell(row=row, column=1, value=f"API (capped at ${cap_refs.get('amount_usd', 1500):,}/day)")
    ws.cell(row=row, column=2, value=f"={comp_refs['api_capped']}"); style_result(ws.cell(row=row, column=2)); fmt_usd(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=f"={comp_refs['total_q']}-{comp_refs['refused']}"); style_formula(ws.cell(row=row, column=3)); fmt_num(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value=f'=CONCATENATE(TEXT({comp_refs["refused"]},"#,##0"), " refused (HTTP 429)")')
    row += 1

    # Row 2: API uncapped
    ws.cell(row=row, column=1, value="API (uncapped — full service)")
    ws.cell(row=row, column=2, value=f"={comp_refs['api_gross']}"); style_result(ws.cell(row=row, column=2)); fmt_usd(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=f"={comp_refs['total_q']}"); style_formula(ws.cell(row=row, column=3)); fmt_num(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value="all queries")
    row += 1

    # Row 3: Self-host Optimistic (η=1.0, 0 FTE, 36-mo amort)
    ws.cell(row=row, column=1, value="Self-host (Optimistic: η=1.0)")
    ws.cell(row=row, column=2, value=f"={comp_refs['self_host_opt']}"); style_result(ws.cell(row=row, column=2)); fmt_usd(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=f"={comp_refs['total_q']}"); style_formula(ws.cell(row=row, column=3)); fmt_num(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value="all queries; quality gap vs frontier (see paper)")
    row += 1

    # Row 4: Self-host Realistic (η=0.75, 0.5 SRE, 12-mo amort)
    ws.cell(row=row, column=1, value="Self-host (Realistic: η=0.75, 0.5 SRE)")
    ws.cell(row=row, column=2, value=f"={comp_refs['self_host_real']}"); style_result(ws.cell(row=row, column=2)); fmt_usd(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=f"={comp_refs['total_q']}"); style_formula(ws.cell(row=row, column=3)); fmt_num(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value="all queries; quality gap vs frontier (see paper)")
    row += 1

    data_end = row - 1

    # Heat-map on the cost column to make the cheapest/most-expensive row obvious
    heatmap_range(ws, f"B{data_start}:B{data_end}")

    # Bar chart comparing the four strategies
    add_bar_chart(
        ws,
        title="Monthly cost by strategy",
        data_ref=f"Output!$B${data_start}:$B${data_end}",
        cats_ref=f"Output!$A${data_start}:$A${data_end}",
        anchor=f"A{row + 2}",
        height=10, width=20,
    )

    # Caveat for the self-host approximation
    row += 16  # leave space for the chart
    ws.cell(row=row, column=1, value="Self-host rows use a single-replica approximation (Eq. 6 simplification).")
    ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True, size=10, color="555555")
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    ws.cell(row=row, column=1, value="The live calculator at calc.ajinkya.ai computes the full peak-tps × η-derate sizing.")
    ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True, size=10, color="555555")
    ws.merge_cells(f"A{row}:D{row}")

    autosize(ws, min_w=26, max_w=80)
    freeze_below(ws, row=5)


def build_sensitivity(ws, w: dict, comp_refs: dict):
    ws.title = "Sensitivity"
    title_banner(
        ws,
        "Sensitivity · ±30% sweep of key inputs",
        "Each row holds the API-capped cost while one input swings by ±30%, ±15%. The headline value lives in the Baseline column.",
        cols=6,
    )

    headers = ["Input swept", "−30%", "−15%", "Baseline", "+15%", "+30%"]
    row = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h); style_header(c, 2)
    row += 1

    data_start = row
    baseline_cell = comp_refs['api_capped']

    # Per-row: a scale factor on the baseline. This is a first-order
    # approximation — for a precise sweep, edit the actual Workload
    # input and re-open the file. But the linear-scale band is correct
    # for traffic-proportional inputs (MAU, sessions/day, q/session).
    sensitivities = [
        ("Total query volume (MAU × sessions × turns)",
            [0.70, 0.85, 1.00, 1.15, 1.30]),
        ("Cache rate baseline (input-cost lever)",
            # Cache rate moves input cost roughly linearly between cached/uncached share
            [1.15, 1.08, 1.00, 0.93, 0.86]),
        ("Per-query input tokens",
            [0.70, 0.85, 1.00, 1.15, 1.30]),
        ("Per-query output tokens",
            # Output is small share of total cost (~10–20%), so the swing is muted
            [0.94, 0.97, 1.00, 1.03, 1.06]),
        ("Daily cap (binding case)",
            # Below baseline, cap constrains; above, no effect because base API rules
            [0.85, 0.93, 1.00, 1.00, 1.00]),
    ]

    for label, factors in sensitivities:
        ws.cell(row=row, column=1, value=label); style_label(ws.cell(row=row, column=1))
        for i, f in enumerate(factors, start=2):
            ws.cell(row=row, column=i, value=f"={baseline_cell}*{f}")
            style_formula(ws.cell(row=row, column=i))
            fmt_usd(ws.cell(row=row, column=i))
        row += 1

    data_end = row - 1

    # Heat-map across the swept range (transposed view: outliers stand out)
    heatmap_range(ws, f"B{data_start}:F{data_end}")

    # Line chart anchored below the data. Build the data range to INCLUDE
    # column A (the swept-input label) so titles_from_data=True pulls
    # each series name from there automatically — no manual SeriesLabel.
    ch_anchor = f"A{row + 2}"
    line_ch = LineChart()
    line_ch.title = "Cost sensitivity"
    line_ch.style = 11
    line_ch.y_axis.title = "API capped cost ($/mo)"
    line_ch.x_axis.title = "% of baseline"
    data = Reference(ws, range_string=f"Sensitivity!$A${data_start}:$F${data_end}")
    line_ch.add_data(data, titles_from_data=True, from_rows=True)
    cats = Reference(ws, range_string=f"Sensitivity!$B$4:$F$4")
    line_ch.set_categories(cats)
    line_ch.height = 10
    line_ch.width = 20
    ws.add_chart(line_ch, ch_anchor)

    row += 18  # space for chart

    # Caveat
    ws.cell(row=row, column=1, value="Note: linear scaling against the baseline; for a precise sweep, edit Workload inputs and re-open.")
    ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True, size=10, color="555555")
    ws.merge_cells(f"A{row}:F{row}")

    autosize(ws, min_w=22, max_w=70)
    freeze_below(ws, row=5)


def build_summary(ws, w: dict, comp_refs: dict, cap_refs: dict):
    """Front-page sheet with the headline answer and key drivers."""
    ws.title = "Summary"
    deployment_name = w["deployment"].get("name", "AI agent deployment")
    title_banner(
        ws,
        f"Summary · {deployment_name}",
        "The procurement-grade one-page view. Everything else in this workbook drives these numbers.",
        cols=4,
    )

    # ----- The headline -----
    row = 4
    ws.cell(row=row, column=1, value="Monthly LLM cost (API, capped)").font = Font(name="Calibri", size=11, color=INK)
    big = ws.cell(row=row, column=2, value=f"={comp_refs['api_capped']}")
    big.font = Font(name="Calibri", size=22, bold=True, color=ACCENT)
    big.fill = PatternFill(start_color=HIGHLIGHT, end_color=HIGHLIGHT, fill_type="solid")
    big.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    big.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    fmt_usd(big)
    ws.row_dimensions[row].height = 36
    row += 2

    # ----- Side-by-side comparison -----
    ws.cell(row=row, column=1, value="Equal-budget comparison"); style_header(ws.cell(row=row, column=1), 2); ws.merge_cells(f"A{row}:D{row}")
    row += 1
    for col, h in enumerate(["Strategy", "Monthly $", "Queries served", "Refused/lost"], start=1):
        c = ws.cell(row=row, column=col, value=h); style_header(c, 3)
    row += 1
    eb_start = row
    ws.cell(row=row, column=1, value="API capped"); ws.cell(row=row, column=2, value=f"={comp_refs['api_capped']}"); fmt_usd(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=f"={comp_refs['total_q']}-{comp_refs['refused']}"); fmt_num(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value=f"={comp_refs['refused']}"); fmt_num(ws.cell(row=row, column=4)); ws.cell(row=row, column=4).font = Font(name="Calibri", color=BAD)
    row += 1
    ws.cell(row=row, column=1, value="API uncapped"); ws.cell(row=row, column=2, value=f"={comp_refs['api_gross']}"); fmt_usd(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=f"={comp_refs['total_q']}"); fmt_num(ws.cell(row=row, column=3)); ws.cell(row=row, column=4, value=0); fmt_num(ws.cell(row=row, column=4))
    row += 1
    ws.cell(row=row, column=1, value="Self-host (Optimistic)"); ws.cell(row=row, column=2, value=f"={comp_refs['self_host_opt']}"); fmt_usd(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=f"={comp_refs['total_q']}"); fmt_num(ws.cell(row=row, column=3)); ws.cell(row=row, column=4, value=0); fmt_num(ws.cell(row=row, column=4))
    row += 1
    ws.cell(row=row, column=1, value="Self-host (Realistic)"); ws.cell(row=row, column=2, value=f"={comp_refs['self_host_real']}"); fmt_usd(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value=f"={comp_refs['total_q']}"); fmt_num(ws.cell(row=row, column=3)); ws.cell(row=row, column=4, value=0); fmt_num(ws.cell(row=row, column=4))
    row += 1
    eb_end = row - 1
    heatmap_range(ws, f"B{eb_start}:B{eb_end}")

    # Bar chart on the summary page (the procurement screenshot)
    add_bar_chart(
        ws,
        title="Equal-budget cost by strategy",
        data_ref=f"Summary!$B${eb_start}:$B${eb_end}",
        cats_ref=f"Summary!$A${eb_start}:$A${eb_end}",
        anchor=f"A{row + 1}",
        height=9, width=18,
    )
    row += 16

    # Key drivers
    ws.cell(row=row, column=1, value="Key drivers"); style_header(ws.cell(row=row, column=1), 2); ws.merge_cells(f"A{row}:D{row}")
    row += 1
    drivers = [
        ("Total monthly queries",         f"={comp_refs['total_q']}", "num"),
        ("Model",                          f"={comp_refs['model_cell']}", "text"),
        ("Cost mode",                      f"={comp_refs['cost_mode']}", "text"),
        ("Daily spend cap",                f"={cap_refs.get('amount_usd', 1500) if isinstance(cap_refs, dict) else 1500}", "usd_simple"),
    ]
    for label, formula, kind in drivers:
        ws.cell(row=row, column=1, value=label); style_label(ws.cell(row=row, column=1))
        cell = ws.cell(row=row, column=2, value=formula); style_formula(cell)
        if kind == "num": fmt_num(cell)
        elif kind == "usd_simple": cell.value = formula.lstrip("="); fmt_usd(cell)
        row += 1

    # Footer pointer to the rest of the workbook
    row += 1
    ws.cell(row=row, column=1, value="See the other sheets to edit inputs (blue cells) or trace the math (formula cells).").font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws.merge_cells(f"A{row}:D{row}")

    autosize(ws, min_w=24, max_w=60)


# ============================================================
# Main
# ============================================================

def generate(workload_path: Path, out_path: Path) -> None:
    with open(workload_path) as f:
        w = json.load(f)

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # README
    readme_ws = wb.create_sheet("README")
    build_readme(readme_ws, w)

    # Workload (inputs)
    workload_ws = wb.create_sheet("Workload")
    workload_refs = build_workload(workload_ws, w)

    # Shapes
    shapes_ws = wb.create_sheet("Shapes")
    shape_rows = build_shapes(shapes_ws, w)

    # Mix
    mix_ws = wb.create_sheet("Mix")
    mix_rows = build_mix(mix_ws, w)

    # Segments
    segments_ws = wb.create_sheet("Segments")
    seg_rows = build_segments(segments_ws, w)

    # Rate Cards
    rates_ws = wb.create_sheet("Rate Cards")
    rate_rows = build_rate_cards(rates_ws, w)

    # Cost Modes
    cost_modes_ws = wb.create_sheet("Cost Modes")
    build_cost_modes(cost_modes_ws)

    # Computation (formulas)
    comp_ws = wb.create_sheet("Computation")
    refs = {
        "anchor_input": workload_refs["anchor_input"],
        "cap": workload_refs["cap"],
        "bot_factor": workload_refs["bot_factor"],
        "shape_rows": shape_rows,
        "mix_rows": mix_rows,
        "seg_rows": seg_rows,
        "rate_rows": rate_rows,
    }
    comp_refs = build_computation(comp_ws, w, refs)

    # Output
    output_ws = wb.create_sheet("Output")
    build_output(output_ws, w, comp_refs, w.get("daily_cap", {"amount_usd": 1500}))

    # Sensitivity
    sens_ws = wb.create_sheet("Sensitivity")
    build_sensitivity(sens_ws, w, comp_refs)

    # Summary (built last so all refs are resolved; moved to front below)
    summary_ws = wb.create_sheet("Summary")
    build_summary(summary_ws, w, comp_refs, w.get("daily_cap", {"amount_usd": 1500}))

    # Reorder so Summary is first; README is naturally second after the move.
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    wb.save(out_path)
    print(f"Wrote {out_path}")


def main():
    ap = argparse.ArgumentParser(description="Generate Excel workbook from a workload spec")
    ap.add_argument("workload", help="Path to workload JSON file")
    ap.add_argument("-o", "--output", help="Output .xlsx path", default=None)
    args = ap.parse_args()

    workload_path = Path(args.workload)
    if not workload_path.exists():
        print(f"ERROR: {workload_path} not found", file=sys.stderr)
        sys.exit(1)
    if args.output:
        out_path = Path(args.output)
    else:
        slug = workload_path.stem
        out_path = workload_path.parent / f"{slug}.xlsx"
    generate(workload_path, out_path)


if __name__ == "__main__":
    main()
